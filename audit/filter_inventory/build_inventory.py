"""Build the BioCirv Phase 1 filter inventory as CSV and XLSX.

Source of truth for the inventory rows. Edit ROWS here, then re-run:

    pixi run -e auditor python audit/filter_inventory/build_inventory.py

Writes filter-inventory.csv and filter-inventory.xlsx alongside this file.
Use push_to_sheet.py to publish the CSV to the shared Google Sheet.

Field definitions follow audit/filter_inventory/filter-inventory-plan.md. "Unknown" is used
where the repository does not document a rationale -- it is not a guess.
"""

from __future__ import annotations

import csv
from pathlib import Path

HERE = Path(__file__).parent

HEADERS = [
    "ID",
    "Rule",
    "Pipeline stage",
    "Shared script",
    "Data affected",
    "Trigger",
    "Effect",
    "Source",
    "Related rules",
    "Questions",
]

# ---------------------------------------------------------------------------
# Stage taxonomy
#
# "Pipeline stage" names the place a reviewer would go looking, using the
# team's own vocabulary. Shared helper modules are NOT their own stage: each
# rule is filed under the stage that consumes it, and the "Shared script"
# column records which helper it comes from, so a diagram can group by helper
# without the stage list fragmenting.
#
# STAGES is ordered; it drives the ordering of the output rows.
# ---------------------------------------------------------------------------
STAGES = [
    "Google Sheets",
    "Extract Scripts",
    "Orchestration (Prefect flows)",
    "Transform Scripts",
    "Load Scripts",
    "Staging + Production (table definitions)",
    "GitHub Action Checks",
    "Materialized Views - data_portal",
    "Materialized Views - ca_biositing",
    "API Endpoints",
    "BioCirV Portal",
    "Stale artifact (not deployed)",
]

STAGE_BY_ID = {
    "F-48": "Google Sheets",
    "F-01": "Extract Scripts", "F-02": "Extract Scripts",
    "F-42": "Orchestration (Prefect flows)", "F-43": "Orchestration (Prefect flows)",
    "F-45": "Orchestration (Prefect flows)", "F-46": "Orchestration (Prefect flows)",
    "F-47": "Orchestration (Prefect flows)",
    "F-03": "Transform Scripts", "F-04": "Transform Scripts", "F-05": "Transform Scripts",
    "F-06": "Transform Scripts", "F-07": "Transform Scripts", "F-08": "Transform Scripts",
    "F-11": "Transform Scripts", "F-44": "Transform Scripts",
    "F-09": "Load Scripts", "F-10": "Load Scripts", "F-14": "Load Scripts",
    # Staging and production share ONE definition tree (datamodels/models/)
    # deployed by the same alembic chain, so these rules cannot be assigned to
    # one environment or the other. Kept as a single combined stage rather than
    # listing every rule twice; see the report for the reasoning.
    "F-12": "Staging + Production (table definitions)",
    "F-13": "Staging + Production (table definitions)",
    "F-15": "Staging + Production (table definitions)",
    # common.py helpers are consumed by BOTH view stacks; filed under
    # data_portal because that is where common.py lives, with the shared-script
    # column marking the reuse.
    **{f"F-{n}": "Materialized Views - data_portal" for n in range(16, 37)},
    "F-37": "Materialized Views - ca_biositing",
    "F-38": "Materialized Views - ca_biositing",
    "F-39": "Stale artifact (not deployed)",
    "F-40": "API Endpoints", "F-41": "API Endpoints",
}

# Stages that are real places filtering could happen but currently hold zero
# rules. Their emptiness is a finding, so they stay in the chart.
EMPTY_STAGES = {
    "GitHub Action Checks":
        "11 workflows gate schema and tests (alembic check, pytest). "
        "Nothing validates data content or row counts.",
    "BioCirV Portal":
        "NOT SCANNED - frontend is an uncheckedout git submodule "
        "(sustainability-software-lab/cal-bioscape-frontend). Gap, not a clean bill of health.",
}

# Which shared helper module each rule's logic actually lives in. Blank means
# the rule is written inline at its own source location.
SHARED_SCRIPT_BY_ID = {
    **{f"F-{n}": "data_portal_views/common.py" for n in range(16, 23)},
    "F-44": "utils/cleaning_functions/cleaning.py",
    # views.py imports the four common.py helpers, so the ca_biositing stack
    # inherits F-16 and F-18 through F-22 rather than restating them.
    "F-38": "data_portal_views/common.py (via views.py import)",
}

# fmt: off
ROWS = [
    # ---------------- GOOGLE SHEETS (ANALYST INPUT) ----------------
    (
        "F-48",
        "An analyst marks a record qc_result = 'fail' in the source spreadsheet",
        "Google Sheets",
        "All 11 analytical record types, at source",
        "Analyst judgement during data entry / QC review, applied in the Google Sheet",
        "Flagged - sets the value that F-17 later acts on. Nothing is removed at this point; the record is still extracted, transformed and stored",
        "Outside the repository. Enters the codebase as the 'qc_result' column, renamed to qc_pass in 10 transform files (calorimetry_record.py:99, compositional_record.py:75, fermentation_record.py:136, gasification_record.py:118, icp_record.py:62, pretreatment_record.py:87, proximate_record.py:74, ultimate_record.py:74, xrd_record.py:98, xrf_record.py:103)",
        "F-17",
        "The single highest-leverage inclusion decision in the system is made by a human in a spreadsheet, with no version control and no documented criteria. What values besides 'pass' and 'fail' occur? Who reviews the decision?",
    ),

    # ---------------- STAGE 1: EXTRACT ----------------
    (
        "F-01",
        "USDA NASS extraction is restricted to California",
        "Extract",
        "usda_census_record, usda_survey_record (all downstream USDA data)",
        "STATE = 'CA' passed to the NASS API request",
        "Removed - non-CA records are never requested and never enter the pipeline",
        "src/ca_biositing/pipeline/ca_biositing/pipeline/etl/extract/usda_census_survey.py:47",
        "F-02",
        "Is CA-only a permanent project boundary or a current-phase scope limit?",
    ),
    (
        "F-02",
        "USDA API response is re-filtered to the 58 valid California county FIPS codes",
        "Extract",
        "usda_census_record, usda_survey_record",
        "county_code not in the seeded CA county FIPS set",
        "Removed - dropped from the dataframe before load; count is logged but rows are not retained",
        "src/ca_biositing/pipeline/ca_biositing/pipeline/etl/extract/usda_census_survey.py:144",
        "F-01, F-25",
        "Defensive de-dup of F-01 or does the API actually return out-of-state rows? The comment says 'in case'.",
    ),

    # ---------------- STAGE 1b: ORCHESTRATION (PREFECT FLOWS) ----------------
    (
        "F-42",
        "Extraction failures are swallowed and the entire source dataset is silently skipped",
        "Orchestration (Prefect flow)",
        "Whole analysis datasets: proximate, ultimate, compositional, icp, xrf, calorimetry, xrd, ftnir",
        "extractor.extract() raises ValueError or IOError",
        "Removed - safe_extract returns None and the dataset is skipped downstream. Logged via logger.exception, but the flow still SUCCEEDS",
        "flows/analysis_records.py:65-67 (safe_extract)",
        "F-43, F-44",
        "One malformed source file silently removes an entire analysis type from that run while the flow reports success. Is this failure mode monitored?",
    ),
    (
        "F-43",
        "Flows abort early and silently when a stage returns an empty dataframe",
        "Orchestration (Prefect flow)",
        "billion_ton, biodiesel_plants, field_sample, aim2_bioconversion, landiq and peer flows",
        "Extract or transform returns None or an empty dataframe",
        "Removed - the flow returns early, load never runs, no exception is raised",
        "flows/billion_ton_etl.py:37,57; flows/biodiesel_plants.py:30,42; flows/aim2_bioconversion.py:53-188 (guards throughout); flows/field_sample_etl.py:68,83",
        "F-42, F-44",
        "'No new data' and 'extraction broke' differ only in log level. Does a zero-row run look identical to a successful no-op to operators?",
    ),
    (
        "F-44",
        "standard_clean returns None for non-dataframe input, dropping that whole source",
        "Shared helper (cleaning.py)",
        "Any source passed through standard_clean, notably the observation transform",
        "Input is not a pandas DataFrame",
        "Removed - returns None and callers 'continue' past it",
        "utils/cleaning_functions/cleaning.py:98-100; consumed at etl/transform/analysis/observation.py:50-52",
        "F-42, F-43",
        "Silent whole-source drop with only an error log. How often does this fire in practice?",
    ),
    (
        "F-45",
        "Gasification archival skips records that are already archived",
        "Orchestration (Prefect flow)",
        "gasification archive uploads",
        "check_record_exists(resource_id, experiment_id) is true AND force_refresh is false",
        "Omitted - record skipped; intended as idempotency, but a changed source sheet is never re-read",
        "flows/gasification_archive_pipeline.py:180-183",
        "F-46, F-13",
        "Existence is keyed on (resource_id, experiment_id), not on source content. An updated GSheet is never picked up unless force_refresh is passed.",
    ),
    (
        "F-46",
        "Gasification records without a source GSheet URL are skipped",
        "Orchestration (Prefect flow)",
        "gasification archive uploads",
        "gsheet_url is empty or missing on the record",
        "Omitted - record skipped with a warning, flow continues",
        "flows/gasification_archive_pipeline.py:185-187",
        "F-45",
        "Unknown - are URL-less gasification records expected, or does this signal incomplete analyst entry?",
    ),
    (
        "F-47",
        "Source worksheets with fewer than four rows are rejected outright",
        "Orchestration (Prefect flow)",
        "gasification archive source Google Sheets",
        "Worksheet has fewer than 4 rows, or the header row (row 4) is empty",
        "Rejected - raises ValueError; unlike F-42 and F-43 this one fails loudly",
        "flows/gasification_archive_pipeline.py:34-42",
        "F-46",
        "Header is assumed to be on row 4. A source sheet with a different layout fails rather than being misread, but the assumption is undocumented.",
    ),

    # ---------------- STAGE 2: TRANSFORM ----------------
    (
        "F-03",
        "Analysis records with a missing record_id are dropped",
        "Transform",
        "calorimetry, compositional, icp, xrf, xrd, proximate, ultimate, fermentation, gasification, pretreatment records",
        "record_id is null after cleaning",
        "Rejected - row dropped in pandas before load",
        "etl/transform/analysis/calorimetry_record.py:121; compositional_record.py:97; icp_record.py:87; fermentation_record.py:165; gasification_record.py:149; pretreatment_record.py:132 (and peers)",
        "F-04, F-12",
        "Are dropped rows counted anywhere? Only some call sites log; no reject table is written.",
    ),
    (
        "F-04",
        "Analysis records whose record_id is a placeholder token are dropped",
        "Transform",
        "fermentation, gasification, pretreatment records",
        "record_id in ['-', 'nan', 'None'/'none', ''] after stripping",
        "Rejected - row dropped in pandas before load",
        "etl/transform/analysis/fermentation_record.py:164; gasification_record.py:148; pretreatment_record.py:131",
        "F-03, F-07",
        "Token lists differ in case ('None' vs 'none'). Does standard_clean lowercase before this runs in every path?",
    ),
    (
        "F-05",
        "Observations are deduplicated on (record_id, record_type, parameter_id, unit_id)",
        "Transform",
        "observation",
        "Duplicate key tuple within or across source dataframes",
        "Discarded - keep='first'; the duplicate count is logged as a warning",
        "etl/transform/analysis/observation.py:137 (per-source), observation.py:156 (cross-source)",
        "F-12",
        "keep='first' means source-file ordering decides which value survives. Is that ordering deterministic?",
    ),
    (
        "F-06",
        "Observations missing record_id, parameter_id, or value are dropped",
        "Transform",
        "observation",
        "Any of the three columns is null",
        "Rejected - row dropped; a warning fires only if ALL rows in a dataframe are dropped",
        "etl/transform/analysis/observation.py:119",
        "F-07, F-19",
        "Partial drops are silent. A source file losing 90% of its rows would not warn.",
    ),
    (
        "F-07",
        "Null XRF and ICP measurement values are coerced to 0, which determines whether they survive F-06",
        "Transform",
        "observation (xrf analysis, icp analysis)",
        "value is null AND analysis_type in ('xrf analysis','icp analysis') AND record_id present; ICP 'y-axial' and 'y-radial' parameters are excepted",
        "Retained - rows that would be dropped by F-06 are instead stored as zero",
        "etl/transform/analysis/observation.py:57-64",
        "F-06",
        "Is a null XRF/ICP reading 'below detection limit' (~0) or 'not measured' (not 0)? Comment attributes this to 'User request' with no rationale. Affects downstream averages.",
    ),
    (
        "F-08",
        "Experiment rows with placeholder or missing names are dropped and deduplicated",
        "Transform",
        "experiment",
        "name in ['-','nan','none',''] or null; then duplicate names",
        "Rejected / discarded - dropped before load",
        "etl/transform/analysis/experiment.py:114-116",
        "F-04",
        "Unknown - are dropped experiments referenced by surviving analysis records, orphaning them?",
    ),
    (
        "F-09",
        "Lookup rows with blank names are dropped",
        "Transform",
        "method, data_source, dataset (county_ag_datasets)",
        "name is null or empty after replacing '', 'nan', 'None', '-' with NA",
        "Rejected - dropped before load; method also deduplicates on name",
        "etl/load/analysis/method.py:24,36; etl/transform/analysis/data_source.py:76,80; etl/transform/analysis/county_ag_datasets.py:61",
        "F-08",
        "Unknown - do records referencing a dropped method lose their method association silently?",
    ),
    (
        "F-10",
        "Residue factor rows without a resource_id are dropped",
        "Transform / Load",
        "residue_factor",
        "resource_id is null",
        "Rejected - dropped before load",
        "etl/load/resource_information/residue_factor.py:56",
        "F-30, F-31",
        "A missing residue factor silently removes a resource from volume estimation entirely (inner join, F-35).",
    ),
    (
        "F-11",
        "Almond NSJV observations are deduplicated on a composite key",
        "Transform",
        "observation (almond NSJV dataset)",
        "Duplicate (record_id | parameter_id | year | resource)",
        "Discarded - drop_duplicates, keep first",
        "etl/transform/analysis/almond_nsjv.py:552",
        "F-05",
        "Dedupe key differs from the general observation key (F-05) - includes year and resource, omits unit_id. Intentional?",
    ),

    # ---------------- STAGE 3: LOAD / DATABASE ----------------
    (
        "F-12",
        "Database unique constraint collapses duplicate observations",
        "Load / Database",
        "observation table",
        "Insert conflicting on (record_id, record_type, parameter_id, unit_id)",
        "Omitted - ON CONFLICT DO UPDATE overwrites the existing row; no new row is created",
        "datamodels/models/general_analysis/observation.py:14-21; etl/load/analysis/observation.py:40",
        "F-05",
        "Third implementation of the same dedup rule (twice in pandas, once in DB). Which one actually fires in production?",
    ),
    (
        "F-13",
        "record_id is unique across analysis record base entities",
        "Load / Database",
        "All aim1/aim2 record tables inheriting BaseEntity",
        "Insert with an existing record_id",
        "Omitted - upsert overwrites; the incoming row does not become a separate record",
        "datamodels/models/base.py:45,65; datamodels/models/external_data/county_ag_report_record.py:9",
        "F-14",
        "Overwrite semantics mean a re-run silently replaces prior values. Is prior state recoverable from etl_run_id lineage?",
    ),
    (
        "F-14",
        "Loads use UPSERT; when there are no updatable columns the row is silently skipped",
        "Load / Database",
        "county_ag_report_record, data_source, and most analysis record loads",
        "ON CONFLICT with an empty update set",
        "Omitted - ON CONFLICT DO NOTHING; success_count is still incremented",
        "etl/load/analysis/county_ag_report_record.py:113-121; etl/load/analysis/data_source.py:80-88",
        "F-13",
        "success_count increments even on DO NOTHING, so load logs overstate rows actually written.",
    ),
    (
        "F-15",
        "One residue factor per (resource_id, factor_type)",
        "Load / Database",
        "residue_factor",
        "Second factor of the same type for a resource",
        "Omitted - unique constraint rejects or upserts over it",
        "datamodels/models/resource_information/residue_factor.py:21",
        "F-10, F-30",
        "Unknown - is a single factor per type scientifically intended, or a modelling simplification?",
    ),

    # ---------------- STAGE 4a: data_portal MATERIALIZED VIEWS ----------------
    (
        "F-16",
        "Seven named resources are excluded from every portal view and the API analysis view",
        "Shared helper (common.py)",
        "All mv_biomass_* views; ca_biositing.analysis_data_view",
        "lower(resource.name) in ['sargassum','#n/a','lab media','alfalfa','almond hulls and shells mix','almond shells and hulls mix','almond woodchips']",
        "Hidden - records remain in base tables but appear in no view",
        "datamodels/data_portal_views/common.py:27-35 (list), :133-137 (filter); applied in every mv_biomass_* file and views.py:290",
        "F-28",
        "Only rationale given is 'exclude problematic records'. Why each name? '#n/a' looks like a data-entry artifact; 'alfalfa' and 'lab media' look deliberate. Should these be one rule or several?",
    ),
    (
        "F-17",
        "Records flagged qc_pass = 'fail' by an analyst are hidden from all views",
        "Shared helper (common.py)",
        "11 record types: compositional, proximate, ultimate, xrf, icp, calorimetry, xrd, ftnir, fermentation, gasification, pretreatment",
        "qc_pass = 'fail' (analyst-supplied 'qc_result' spreadsheet column)",
        "Hidden - never removed from base tables; visible to anyone with direct DB access",
        "common.py:65-75 (central); restated at mv_biomass_composition.py:65,99,109; mv_biomass_fermentation.py:85,144; mv_biomass_gasification.py:52; mv_biomass_sample_stats.py:38; views.py:291",
        "F-18, F-21, F-22",
        "Code is != 'fail' but docstrings at mv_biomass_composition.py:9 and mv_biomass_sample_stats.py:6 claim == 'pass'. NULL and any other value are ADMITTED. How many records have qc_pass outside {pass, fail}?",
    ),
    (
        "F-18",
        "Ultimate analysis is restricted to five whitelisted parameters",
        "Shared helper (common.py)",
        "mv_biomass_composition, ca_biositing.analysis_data_view",
        "analysis_type is ultimate AND lower(parameter) not in ['carbon','nitrogen','oxygen','sulfur','hydrogen']",
        "Omitted - observation excluded from the view",
        "common.py:38 (list), :140-158 (filter); mv_biomass_composition.py:66; views.py:292",
        "F-19",
        "Unknown - why these five? Are other ultimate parameters (e.g. chlorine, moisture) measured and intentionally suppressed, or never collected?",
    ),
    (
        "F-19",
        "Ultimate analysis values above 100 are excluded",
        "Shared helper (common.py)",
        "mv_biomass_composition, ca_biositing.analysis_data_view",
        "analysis_type is ultimate AND value > 100",
        "Omitted - observation excluded (values are percentages)",
        "common.py:151-157 (helper, when value_col supplied); mv_biomass_composition.py:68-71 (inline duplicate); views.py:292",
        "F-18",
        "Implemented twice with different analysis_type matching: helper checks 3 string variants, inline checks only 'ultimate'. Agree today; fragile if the literal changes.",
    ),
    (
        "F-20",
        "ICP observations not reported in ppm are excluded",
        "Shared helper (common.py)",
        "mv_biomass_composition, ca_biositing.analysis_data_view",
        "analysis_type is icp/icp-oes/icp-ms AND lower(unit) != 'ppm'",
        "Omitted - observation excluded from the view",
        "common.py:161-166; mv_biomass_composition.py:67; views.py:293",
        "F-23",
        "Records in mg/kg (numerically equal to ppm) are excluded rather than converted. Intended?",
    ),
    (
        "F-21",
        "Proximate experiments whose component sum falls outside 95-105 are excluded",
        "Shared helper (common.py)",
        "mv_biomass_composition, ca_biositing.analysis_data_view",
        "moisture + ash solids + volatile solids (or 100 - fixed carbon) is not 0 and not within [95,105], per (resource, experiment, analysis_type)",
        "Omitted - the whole experiment group is excluded, not just the offending observation",
        "mv_biomass_composition.py:118-148,178-189; common.py:169-220; views.py:294-305",
        "F-22, F-17",
        "Sum of exactly 0 is treated as 'no data' and admitted. Can a genuine all-zero measurement set be wrongly admitted this way?",
    ),
    (
        "F-22",
        "Compositional experiments whose component sum falls outside 40-105 are excluded",
        "Shared helper (common.py)",
        "mv_biomass_composition, ca_biositing.analysis_data_view",
        "glucan + xylan + lignin (or lignin+) is not 0 and not within [40,105], per (resource, experiment, analysis_type)",
        "Omitted - the whole experiment group is excluded",
        "mv_biomass_composition.py:130-137,190-200; common.py:204-215; views.py:306-316",
        "F-21",
        "Why 40 as a lower bound where proximate uses 95? Unknown - no documented rationale. lignin vs lignin+ handled differently here than in mv_biomass_search.py:66-76.",
    ),
    (
        "F-23",
        "ICP experiments containing any value above 500,000 ppm are excluded",
        "View (data_portal)",
        "mv_biomass_composition only",
        "max ICP value in ppm > 500000 for a (resource, experiment) group",
        "Omitted - the whole experiment group is excluded",
        "mv_biomass_composition.py:138-143,201-208",
        "F-20",
        "Present in the portal view but ABSENT from ca_biositing.analysis_data_view. Same experiment is hidden in the portal and visible via the API. Intended?",
    ),
    (
        "F-24",
        "Observation record_type must be one of eleven analytical types",
        "View (data_portal)",
        "analysis_metrics subquery feeding mv_biomass_search",
        "record_type not in the eleven-value whitelist",
        "Omitted - observation excluded from metric aggregation",
        "common.py:54-59",
        "F-17, F-40",
        "Whitelist rather than blacklist: a newly added analysis type is silently invisible until added here.",
    ),
    (
        "F-25",
        "Volume estimates are restricted to San Joaquin, Stanislaus and Merced counties",
        "View (data_portal)",
        "mv_biomass_volume_estimate (all five paths), and mv_biomass_search volume columns downstream",
        "lower(place.county_name) not in ['san joaquin','stanislaus','merced']",
        "Hidden - the other 55 CA counties are extracted, stored and indexed, then filtered out at the view layer",
        "mv_biomass_volume_estimate.py:87, :182, :257, :351, :405 (five separate literals)",
        "F-01, F-02",
        "Written five times as a bare literal, not a named constant like EXCLUDED_RESOURCES. Is NSJV-only a permanent scope or a pilot boundary? Added in migration 0020.",
    ),
    (
        "F-26",
        "Volume estimation ignores data years before 2017",
        "View (data_portal)",
        "mv_biomass_volume_estimate (all paths), mv_usda_county_production",
        "usda_census_record.year < 2017 or county_ag_report_record.data_year < 2017",
        "Omitted - older records are stored but excluded from volume views",
        "mv_biomass_volume_estimate.py:85,:178,:255,:349,:403; mv_usda_county_production.py:74",
        "F-25",
        "Unknown - why 2017? Extract sets YEAR = None so all years are pulled and stored.",
    ),
    (
        "F-27",
        "'almond meats' no longer contributes to almond hulls volume",
        "View (data_portal)",
        "mv_biomass_volume_estimate Path C (commodity_direct)",
        "The 'almond meats' -> 'almond hulls' pair was removed from the commodity name map",
        "Removed - almond meats production no longer counts toward almond hulls",
        "alembic/versions/0021_remove_almond_meats_from_volume_estimate.py; source now at mv_biomass_volume_estimate.py:199-206",
        "F-28, F-16",
        "Python source and migration 0021 agree. Was the prior mapping double-counting, or was almond meats deemed out of scope?",
    ),
    (
        "F-28",
        "Only six hardcoded commodity name pairs receive commodity-direct volumes",
        "View (data_portal)",
        "mv_biomass_volume_estimate Path C",
        "lower(primary_ag_product.name) not present in the six-row literal map",
        "Omitted - resource gets no Path C volume, silently",
        "mv_biomass_volume_estimate.py:199-214",
        "F-27, F-16, F-29",
        "String-literal matching on names. A renamed primary_ag_product silently zeroes volumes with no error.",
    ),
    (
        "F-29",
        "Path E excludes resources that already have county ag report production",
        "View (data_portal)",
        "mv_biomass_volume_estimate Path E (census_production_based)",
        "Resource has any linked county_ag_report_record via primary_ag_product",
        "Omitted - anti-join, stated purpose is to avoid double counting",
        "mv_biomass_volume_estimate.py:406-412",
        "F-25, F-28",
        "The only anti-double-counting guard in the view. Paths A and C both read county_ag_report_record with no equivalent guard. Header comment at :428 claims 'precedence logic' that does not exist.",
    ),
    (
        "F-30",
        "Volume paths are gated on residue factor type",
        "View (data_portal)",
        "mv_biomass_volume_estimate Paths A, B, D, E",
        "factor_type must equal 'weight' (A, E), not equal 'area' (B), or equal 'area' (D)",
        "Omitted - resource excluded from the non-matching path",
        "mv_biomass_volume_estimate.py:84 (A), :179 (B), :402 (E); Path D via join",
        "F-10, F-15, F-31",
        "A resource with no residue factor of the right type gets no volume at all. Is that distinguishable from a genuine zero in the portal?",
    ),
    (
        "F-31",
        "Path B requires a non-null prune_trim_yield",
        "View (data_portal)",
        "mv_biomass_volume_estimate Path B (census_based)",
        "residue_factor.prune_trim_yield is null",
        "Omitted - resource excluded from Path B",
        "mv_biomass_volume_estimate.py:178",
        "F-30, F-10",
        "Comment at :288 notes grape vine prunings (resource 32) is excluded for lacking a usda_commodity_map entry - a related silent exclusion.",
    ),
    (
        "F-32",
        "The aggregate 'NSJV' geoid is excluded from production totals",
        "View (data_portal)",
        "mv_biomass_search production volume aggregation",
        "resource_production_record.geoid = 'NSJV'",
        "Omitted - excluded from both the latest-year subquery and the sum",
        "mv_biomass_search.py:212, :228",
        "F-25",
        "Comment calls NSJV 'an outlier and not mappable to a single geoid'. Are NSJV-only records therefore invisible in every volume output?",
    ),
    (
        "F-33",
        "Fermentation records failing sugar-consumption consistency are excluded",
        "View (data_portal)",
        "mv_biomass_fermentation",
        "abs(avg_sugar_cons - ((sugart0 - sugarteof)/sugart0)*100) > 100; bypassed when any input is null or sugart0 = 0",
        "Omitted - record excluded from the view",
        "mv_biomass_fermentation.py:146-158",
        "F-34, F-17",
        "A 100% tolerance admits nearly everything. Comment says '~100% tolerance'. Is the threshold intentional or a placeholder?",
    ),
    (
        "F-34",
        "Fermentation yield values outside 0-105 are excluded",
        "View (data_portal)",
        "mv_biomass_fermentation",
        "parameter name matches '%yield%' AND avg(value) outside [0,105]",
        "Omitted - HAVING clause drops the grouped row",
        "mv_biomass_fermentation.py:161-168",
        "F-33, F-21",
        "Applied post-aggregation, so an out-of-range individual observation can be masked by averaging. Deliberate?",
    ),
    (
        "F-35",
        "Inner joins to Resource, Place and ResidueFactor drop unmatched records",
        "View (data_portal)",
        "mv_biomass_composition, mv_biomass_volume_estimate (all paths), mv_biomass_county_production, mv_billion_ton_county_production, mv_usda_county_production, mv_biomass_gasification, mv_biomass_fermentation",
        "No matching resource, place (geoid) or residue factor row",
        "Removed by join - record silently absent from the view",
        "mv_biomass_composition.py:165; mv_biomass_volume_estimate.py:79,166,247,338; mv_biomass_county_production.py:40,42; mv_usda_county_production.py:68-72",
        "F-10, F-30",
        "Unmatched geoids are invisible with no warning. How many base records fail these joins? Needs a live row-count.",
    ),
    (
        "F-36",
        "Pricing and end-use views restrict record types and parameter names",
        "View (data_portal)",
        "mv_biomass_pricing, mv_biomass_end_uses",
        "record_type not in ['usda_market_record','resource_price_record'] (pricing) or != 'resource_end_use_record' (end uses); parameter not in the price/end-use name whitelists",
        "Omitted - observation excluded from the view",
        "mv_biomass_pricing.py:53-56; mv_biomass_end_uses.py:101,120-123",
        "F-24",
        "Parameter-name whitelists matched as lowercase literals. A renamed parameter silently disappears.",
    ),

    # ---------------- STAGE 4b: ca_biositing ANALYTICAL VIEWS (API) ----------------
    (
        "F-37",
        "analysis_data_view excludes USDA census and survey observations",
        "View (ca_biositing)",
        "ca_biositing.analysis_data_view, and analysis_average_view derived from it",
        "lower(record_type) in ('usda_census_record','usda_survey_record')",
        "Omitted - routed to the separate usda_census_view / usda_survey_view instead",
        "views.py:245-249; resources/sql/create_analytical_views.sql:77",
        "F-38, F-24",
        "The one rule both artifacts agree on. Consumers must query three views to see all observations.",
    ),
    (
        "F-38",
        "The deployed analysis_data_view DOES apply the full QC filter set (resolved)",
        "View (ca_biositing)",
        "ca_biositing.analysis_data_view -> all REST API analysis endpoints",
        "Same triggers as F-16, F-17, F-18, F-19, F-20, F-21, F-22",
        "Hidden - RESOLVED. The deployed SQL in migration 0021 contains qc_pass (21x), the excluded-resource names, the ultimate and ICP filters, and both sum bounds",
        "alembic/versions/0021_remove_almond_meats_from_volume_estimate.py (compiled from views.py:288-321). NOT from resources/sql/create_analytical_views.sql, which nothing in the repo executes",
        "F-16, F-17, F-18, F-19, F-20, F-21, F-22, F-23, F-39",
        "RESOLVED by inspecting the deployed migration SQL. The API is NOT serving QC-failed records. Remaining question is only whether staging has migrations applied through 0021.",
    ),
    (
        "F-39",
        "Inner join to unit drops observations with no unit -- in an orphaned artifact only",
        "Stale artifact (not deployed)",
        "resources/sql/create_analytical_views.sql only; NOT the deployed views",
        "observation.unit_id is null",
        "No live effect - no code path in the repository executes this file",
        "resources/sql/create_analytical_views.sql:53, :96, :117",
        "F-38",
        "The file is orphaned: grep finds zero references outside this inventory. It also lacks every QC filter the deployed view has. Recommend deleting or regenerating it - as written it misleads readers about live behaviour.",
    ),

    # ---------------- STAGE 5: API ----------------
    (
        "F-40",
        "API responses are capped at 100 records per request",
        "API",
        "All paginated feedstock endpoints",
        "limit query parameter, default 100, hard maximum 100",
        "Hidden - records beyond the page are withheld until the client requests the next offset",
        "src/ca_biositing/webservice/ca_biositing/webservice/dependencies.py:31-34",
        "F-41",
        "Offset pagination with no total-count field found. A client that does not paginate sees a silently truncated dataset.",
    ),
    (
        "F-41",
        "Facet endpoints omit rows with null facet values",
        "API",
        "Resource, geoid and parameter listing endpoints",
        "view column is_not(None) filters on resource / geoid / parameter / usda_crop",
        "Omitted - excluded from facet lists, so the values are unreachable through the UI filters",
        "services/analysis_service.py:213,229,241; services/usda_census_service.py:345,357",
        "F-35, F-40",
        "Records with a null geoid exist in the view but cannot be reached via geoid-based navigation. Are they reachable another way?",
    ),
]
# fmt: on


def build_rows() -> list[list[str]]:
    """Apply the stage taxonomy and shared-script column, then sort by stage.

    ROWS carries the original per-rule stage string in position 2; it is
    replaced from STAGE_BY_ID so the taxonomy lives in one place. Any rule ID
    missing from the map is a bug and raises rather than silently mis-filing.
    """
    stage_order = {name: i for i, name in enumerate(STAGES)}
    out = []
    for row in ROWS:
        rule_id = row[0]
        if rule_id not in STAGE_BY_ID:
            raise KeyError(f"{rule_id} has no stage in STAGE_BY_ID")
        stage = STAGE_BY_ID[rule_id]
        shared = SHARED_SCRIPT_BY_ID.get(rule_id, "")
        out.append([row[0], row[1], stage, shared, *row[3:]])
    out.sort(key=lambda r: (stage_order[r[2]], r[0]))
    return out


def main() -> None:
    rows = build_rows()

    csv_path = HERE / "filter-inventory.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS)
        writer.writerows(rows)
    print(f"Wrote {csv_path} ({len(rows)} rules)")

    from collections import Counter
    print("\nRules by stage:")
    counts = Counter(r[2] for r in rows)
    for stage in STAGES:
        n = counts.get(stage, 0)
        note = f"   <- {EMPTY_STAGES[stage]}" if stage in EMPTY_STAGES else ""
        print(f"  {n:3}  {stage}{note}")
    shared = Counter(r[3] for r in rows if r[3])
    print("\nRules originating in a shared script:")
    for script, n in shared.most_common():
        print(f"  {n:3}  {script}")
    print()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not available - skipping XLSX. CSV is complete.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Filter Inventory"

    ws.append(HEADERS)
    for row in rows:
        ws.append(list(row))

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {"ID": 7, "Rule": 52, "Pipeline stage": 26, "Shared script": 26,
              "Data affected": 40, "Trigger": 46, "Effect": 40, "Source": 56,
              "Related rules": 18, "Questions": 60}
    for idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths[header]

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    xlsx_path = HERE / "filter-inventory.xlsx"
    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path}")


if __name__ == "__main__":
    main()
